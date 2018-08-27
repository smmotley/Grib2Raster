'''
Version 0.1
Created on February 1, 2018
@author: Shane Motley

@Purpose: 1) Pull data from grib file into numpy array
          2) Convert data from a given coordinate system into lat/lng.
          3) Given a polygon (e.g. river basin), determine which points in the raster file are in the polygon
          4) Create a numpy mask and set points outside the polygon to zero and perform necessary calculations.
          5) Create any plots needed.

@Usage: User can call program and get the forecast for the following calls:
        -l: for a single point (lat / lon). EX: -l 34.5 -104.3
        -hr: hourly (will interpolate forecast hours to generate hourly values)
        -help: help menu


@VersionHistory:
    v0.1: 2/01/18 Beta version
'''

from osgeo import osr
from osgeo.gdalnumeric import *
from osgeo.gdalconst import *
from pyproj import Proj
from shapely.geometry import Point, Polygon
from PIL import Image, ImageFont, ImageDraw
import datetime
from mpl_toolkits.basemap import Basemap
import os, sys
import numpy as np
import pytz
import matplotlib.pyplot as plt
from mpl_toolkits.axes_grid1.inset_locator import inset_axes
from mpl_toolkits.axes_grid1.inset_locator import mark_inset
import time
import seaborn as sb
from openpyxl import load_workbook
import pandas as pd
import fiona
import argparse
import SNODAS_Download as sd

class Grib():
    '''
    class object for holding the grib data
    '''

    def __init__(self):
        self.model = ""
        self.date = np.array([])
        self.basin = ""
        self.level = ""
        self.acc = "" #0-3 hr accumulation or 1hr forecast
        self.gribAll = ""
        self.units = ""
        self.ptVal = []
        self.displayunits = ""
        self.data = np.array([])
        self.elevation_data = np.array([])
        self.basinTotal = np.array([])
        self.bbox = []

#you will want to remove "type", this is just for testing.
def handle_args(argv):

    parser = argparse.ArgumentParser()
    parser.add_argument('model',
                    nargs="*",
                    default='snodas')

    parser.add_argument('date',
                    nargs="*",
                    default=time.strftime("%Y%m%d"))  # This will be today's date in yyyymmdd format)

    parser.add_argument('-d', '--date2',
                        dest='date2', required=False,
                        default=None)  # This will be today's date in yyyymmdd format)

    parser.add_argument('-b','--basin',
                        dest='basin', default='French_Meadows',
                        required=False,
                        help='The basin to calculate Total SWE. Options inlude:\n'
                             'Hell_Hole, French_Meadows, or MFP')

    parser.add_argument('-u', '--displayunits',
                        dest='displayunits', default='US',
                        required=False,
                        help='Show Units In US or SI')

    parser.add_argument('-l', '--level',
                        dest='level', default='0-SFC',
                        required=False,
                        help='Variable level EX: -l 0-SFC ')

    parser.add_argument('-m', '--map',
                        dest='map', default=True, required=False,
                        help='use this option if you want to output a map')

    parser.add_argument('-p', '--plot',
                        dest='plot', default=True, required=False,
                        help='use this option if you want to output a plot')

    args = parser.parse_args()

    return args

def main():
    global inputArgs, grib, dir_path                          #Make our global vars: grib is the object that will hold our Grib Class.
    dir_path = os.path.dirname(os.path.realpath(__file__))
    comparison_days =[0,-7]
    inputArgs = handle_args(sys.argv)               #All input arguments if run on the command line.
    for deltaDay in comparison_days:
        if deltaDay == 0:
            date2 = None
        else:
            date2 = ((datetime.datetime.now(pytz.timezone('US/Pacific'))) + datetime.timedelta(days=deltaDay)).strftime(
            "%Y%m%d")

        ##############
        # Debugging
        #inputArgs.date = '20180327'
        inputArgs.date = time.strftime("%Y%m%d")
        inputArgs.date2 = date2 #Comment this out for just one date
        inputArgs.map = True  # Make the map and save png to folder.
        findValueAtPoint = False  # Find all the values at specific lat/lng points within an excel file.
        #################

        grib = Grib()                                   #Assign variable to the Grib Class.
        grib.model = inputArgs.model                    #Our model will always be "snodas" for this program
        grib.displayunits = inputArgs.displayunits
        grib.basin = inputArgs.basin                    # Basin can be "French_Meadows", "Hell_Hole", or "MFP", this gets shapefile

        # Bounding box will clip the raster to focus in on a region of interest (e.g. CA) This makes the raster MUCH smaller
        # and easier to work with. See gdal.Open -> gdal.Translate below for where this is acutally used.
        grib.bbox = [-125.0,50.0,-115.0,30.0]           #[upper left lon, upper left lat, lower left lon, lower left lat]
        grib = get_snowdas(grib, inputArgs.date)                        #Get the snodas file and save data into the object variable grib
        #pngFile = makePNG()
        #Any reprojections of grib.gribAll have already been done in get_snowdas.
        #The original projection of snodas is EPSG:4326 (lat/lng), so it has been changed to EPSG:3875 (x/y) in get_snodas
        projInfo = grib.gribAll.GetProjection()

        geoinformation = grib.gribAll.GetGeoTransform() #Get the geoinformation from the grib file.

        xres = geoinformation[1]
        yres = geoinformation[5]
        xmin = geoinformation[0]
        xmax = geoinformation[0] + (xres * grib.gribAll.RasterXSize)
        ymin = geoinformation[3] + (yres * grib.gribAll.RasterYSize)
        ymax = geoinformation[3]

        spatialRef = osr.SpatialReference()
        spatialRef.ImportFromWkt(projInfo)
        spatialRefProj = spatialRef.ExportToProj4()

        # create a grid of xy (or lat/lng) coordinates in the original projection
        xy_source = np.mgrid[xmin:xmax:xres, ymax:ymin:yres]
        xx, yy = xy_source

        # A numpy grid of all the x/y into lat/lng
        # This will convert your projection to lat/lng (it's this simple).
        lons, lats = Proj(spatialRefProj)(xx, yy, inverse=True)


        # Find the center point of each grid box.
        # This says move over 1/2 a grid box in the x direction and move down (since yres is -) in the
        # y direction. Also, the +yres (remember, yres is -) is saying the starting point of this array will
        # trim off the y direction by one row (since it's shifted off the grid)
        xy_source_centerPt = np.mgrid[xmin + (xres / 2):xmax:xres, ymax + (yres / 2):ymin:yres]
        xxC, yyC = xy_source_centerPt

        lons_centerPt, lats_centerPt = Proj(spatialRefProj)(xxC, yyC, inverse=True)

        mask = createMask(xxC, yyC, spatialRefProj)
        grib.basinTotal = calculateBasin(mask, grib, xres, yres)

        # Calculate the difference between two rasters
        if inputArgs.date2 != None:
            grib.basinTotal[0] = compareDates(mask, grib, xres, yres)[0]

        if grib.basin == 'Hell_Hole': #Part of this basin is SMUD's teritory, so remove 92% of water in this basin
            grib.basin = 'Hell_Hole_SMUD' #This is just to get the correct directory structure
            submask = createMask(xxC, yyC, spatialRefProj)
            smudBasinTotal = calculateBasin(submask, grib, xres, yres)
            print("Extracting 92% of the SWE values from SMUD Basin...\n" + "Current Basin Total: " + str(grib.basinTotal[0]))
            grib.basinTotal[0] = grib.basinTotal[0] - (0.92*smudBasinTotal[0])
            print("Smud Total: "+str(smudBasinTotal[0])+"\n New Total: "+str(grib.basinTotal[0]))
            grib.basin = 'Hell_Hole' #reset back

        #Need to do this after Heel_Hole's data has been manipulated (to account for SMUD)
        elevation_bins = calculateByElevation(mask, grib, xres, yres)

        #Send data for writing to Excel File
        if deltaDay == 0:
            excel_output(elevation_bins)

        if inputArgs.plot == True:
            makePlot(elevation_bins, deltaDay)
        print(elevation_bins)

        print(inputArgs.date," Basin Total: ", grib.basinTotal[0])

        #findValue will return a dataframe with SWE values at various lat/lng points.
        df_ptVal = None
        if findValueAtPoint == True:
            df_ptVal = findPointValue(spatialRefProj, xy_source)

        if inputArgs.map == True:
            fig = plt.figure()
            ax = fig.add_subplot(111)
            m = Basemap(llcrnrlon=-122.8, llcrnrlat=37.3,
                        urcrnrlon=-119.0, urcrnrlat=40.3, ax=ax)

            m.arcgisimage(service='ESRI_Imagery_World_2D', xpixels=2000, verbose=True)
            #m.arcgisimage(service='World_Shaded_Relief', xpixels=2000, verbose=True)

            #For inset
            # loc =>'upper right': 1,
            # 'upper left': 2,
            # 'lower left': 3,
            # 'lower right': 4,
            # 'right': 5,
            # 'center left': 6,
            # 'center right': 7,
            # 'lower center': 8,
            # 'upper center': 9,
            # 'center': 10
            axin = inset_axes(m.ax, width="40%", height="40%", loc=8)

            m2 = Basemap(llcrnrlon=-120.7, llcrnrlat=38.7,
                         urcrnrlon=-120.1, urcrnrlat=39.3, ax=axin)

            m2.arcgisimage(service='ESRI_Imagery_World_2D', xpixels=2000, verbose=True)
            mark_inset(ax, axin, loc1=2, loc2=4, fc="none", ec="0.5")

            ###################################DEBUGGING AREA###############################################################
            # Debugging: Test to prove a given lat/lng pair is accessing the correct grid box:

            #*********TEST 1: Test for center points
            #grib.data[0,0] = 15 #increase the variable by some arbitrary amount so it stands out.
            #xpts, ypts = m(lons_centerPt[0,0],lats_centerPt[0,0]) #This should be in the dead center of grid[0,0]
            #m.plot(xpts,ypts, 'ro')

            #*********TEST 2: Test for first grid box
            # Test to see a if the point at [x,y] is in the upper right corner of the cell (it better be!)
            #xpts, ypts = m(lons[0, 0], lats[0, 0])  # should be in upper right corner of cell
            #m.plot(xpts, ypts, 'bo')

            # *********TEST 3: Test for first grid box
            # Test to see the location of center points of each grid in polygon
            # To make this work, uncomment the variables in def create_mask
            #debug_Xpoly_center_pts, debug_Ypoly_center_pts = m(debugCenterX, debugCenterY)
            #m.plot(debug_Xpoly_center_pts, debug_Ypoly_center_pts, 'bo')

            # *********TEST 4: Test grid box size (In lat lng coords)
            # This is for use in a Basemap projection with lat/lon (e.g. EPSG:4326)
            #testX = np.array([[-120.1, -120.1], [-120.10833, -120.10833]])
            #testY = np.array([[39.0, 39.00833], [39.0, 39.00833]])
            # testVal = np.array([[4,4],[4,4]])

            # For use in basemap projection with x/y (e.g. espg:3857. In m=basemap just include the argument projection='merc')
            # testX = np.array([[500975, 500975], [(500975 + 1172), (500975 + 1172)]])
            # testY = np.array([[502363, (502363 + 1172)], [502363, (502363 + 1172)]])
            #testVal = np.array([[18, 18], [18, 18]])
            #im1 = m.pcolormesh(testX, testY, testVal, cmap=plt.cm.jet, vmin=0.1, vmax=10, latlon=False, alpha=0.5)

            # Test to see all points
            # xtest, ytest = m(lons,lats)
            # m.plot(xtest,ytest, 'bo')
            ################################################################################################################

            hr = 0
            makeMap(lons, lats, hr, m, m2,df_ptVal, deltaDay)

    return


def get_snowdas(gribObj,date):
    baseFolder = os.path.join(dir_path,'grib_files')
    snowdas_dir = os.path.join(baseFolder,date,grib.model)
    fyear, fmonth, fday = '1970', '1', '31'  # Filling with random data
    if not os.path.exists(snowdas_dir):
        sd.main('snodas',date)
    if os.path.exists and os.listdir(snowdas_dir)==[]: #it's empty
        sd.main('snodas', date)
    for file in os.listdir(snowdas_dir):
        if file.endswith('.Hdr'):
            gribObj.gribAll = gdal.Open(os.path.join(snowdas_dir,file), GA_ReadOnly)
            #<--Extract Date Info-->
            f = open(snowdas_dir + './' + file, 'r')
            year_str = 'Created year'
            month_str = 'Created month'
            day_str = 'Created day'
            for line in f:
                if year_str in line:
                    fyear = line[-5:]  # get last 5 charaters (one is a space)
                if month_str in line:
                    fmonth = line[-3:]
                if day_str in line:
                    fday = line[-3:]
            #<--Done with Date Info-->
            f.close()
            #Check to see if we are comparing two dates, if we are and this is the second date, grib.date will have a value
            if gribObj.date != None:
                gribObj.date2 = datetime.datetime(year=int(fyear), month=int(fmonth),day=int(fday))  # This will be the date in yyyymmdd format
            else:
                gribObj.date =  datetime.datetime(year=int(fyear), month=int(fmonth),day=int(fday))  # This will be the date in yyyymmdd format
    #Notes: This next section is important because:
    #       1) We are quickly reducing the size of raster using the "projWin=" parameter.
    #       2) We are transforming the SNODAS grid from latlon (EPSG:4326) to XY (EPSG:3857)
    # We MUST transform this into XY coordinates because we are making calculations on the rasters based off of
    # meters (not decimal degrees). For example, once we use gdal.Warp, it will reproject it into xy coordinates, which
    # we can then use to find the x,y resolution of each grid box in meters. Now, we will know that for every raster
    # grid cell, we can calculate any type of Volume calculation we need, like 100 mm of rainfall over a 1,000 x 1,000 m
    # grid will give us ~8 acre feet.
    #Clip the raster to a given bounding box (makes the raster much easier to work with)
    gribObj.gribAll = gdal.Translate('/vsimem/temp.dat', gribObj.gribAll, projWin=gribObj.bbox)
    projInfo = grib.gribAll.GetProjection()
    try:
        spatialRef = osr.SpatialReference(wkt=projInfo)
        cord_sys = spatialRef.GetAttrValue("GEOGCS|AUTHORITY",1)
    except:
        print ("NO COORDINATE SYSTEM FOUND! Transforming to XY")
        cord_sys = '4326'
    if cord_sys != '3875':
        gribObj.gribAll = gdal.Warp('/vsimem/temp.dat', gribObj.gribAll,dstSRS='EPSG:3857')  # If you wanted to put it into x/y coords
        print("Successfully Reprojected Coordinate System From Lat/Lng to X/Y")
    band = gribObj.gribAll.GetRasterBand(1)
    data = BandReadAsArray(band)
    data[data == -9999] = 0
    gribObj.data = data
    gribObj.units = '[kg/(m^2)]' #the units are in mm, or kg/m^2
    return gribObj

def createMask(xxC,yyC,spatialRefProj):

    # Get the lat / lon of each grid box's center point.
    lons_centerPt, lats_centerPt = Proj(spatialRefProj)(xxC, yyC, inverse=True)

    # given a bunch of lat/lons in the geotiff, we can get the polygon points.
    sf = fiona.open(dir_path+'/Shapefiles/'+grib.basin+'/'+grib.basin+'.geojson')
    geoms = [feature["geometry"] for feature in sf]
    poly = Polygon(geoms[0]['coordinates'][0])  # For a simple square, this would be 4 points, but it can be thousands of points for other objects.
    polyMinLon, polyMinLat, polyMaxLon, polyMaxLat = sf.bounds  # Get the bounds to speed up the process of generating a mask.

    # create a 1D numpy mask filled with zeros that is the exact same size as the lat/lon array from our projection.
    mask = np.zeros(lons_centerPt.flatten().shape)

    # Debugging: FOR CENTER POINT OF GRID BOX
    # These are test variables to see where the center points are (plot with basemaps to prove to yourself they're in the right spot).
    global debugCenterX, debugCenterY
    debugCenterX = np.zeros(lats_centerPt.flatten().shape)
    debugCenterY = np.zeros(lons_centerPt.flatten().shape)

    # Create Mask by checking whether points in the raster are within the bounds of the polygon. Instead of checking
    # every single point in the raster, just focus on points within the max/min bounds of the polygon (it's slow as hell
    # if you don't do that).
    i = 0  # counter
    for xp, yp in zip(lons_centerPt.flatten(), lats_centerPt.flatten()):
        if ((polyMinLon <= xp <= polyMaxLon) and (polyMinLat <= yp <= polyMaxLat)):
            mask[i] = (Point(xp, yp).within(poly))
            # Debugging FOR CENTER POINT OF GRID BOX: If you want to visualize the center point
            #       of each grid box that's found in the polygon,
            #       include this below and then you can put a dot (via m.plot)
            if (Point(xp, yp).within(poly)):
                debugCenterX[i] = xp
                debugCenterY[i] = yp
        i += 1

    mask = np.reshape(mask, (xxC.shape))
    return mask

def calculateBasin(mask, gribObj, xres, yres):
    basinTotal = 0
    if gribObj.units == '[kg/(m^2)]':
        value_mm = gribObj.data.copy()
        value_m = value_mm * 0.001  # mm to m.
        value_inches = gribObj.data.copy() * 0.03937  # mm to inches.
        #value_Cubed = totalPrecip_mm * abs(xres) * abs(yres)  # change each grid box to total precip in cubic meters
        value_AF = (value_m * abs(xres) * abs(yres))/ 1233.48  # 1 AF = 1233.48 m^3
        basinSWE_in = (np.average(value_inches))
        basinTotal = [np.sum(mask * value_AF.T)]
    return basinTotal

def calculateByElevation(mask, gribObj, xres, yres):
   #If elevation data is empty, fill it up (this obviously won't change, so just do it once)
    if not gribObj.elevation_data:
        elevationRaster() #Create the elevation raster and store data in grib.elevation_data
    elevationRange = np.arange(0,10000,500) #np.arrange(start,stop,step)
    df = pd.DataFrame(columns=['ElevRange','TotalAF','AveSWE'])
    if gribObj.units == '[kg/(m^2)]':
        value_mm = gribObj.data.copy()
        value_m = value_mm * 0.001  # mm to m.
        value_inches = gribObj.data * 0.03937  # mm to inches.
        value_AF = (value_m * abs(xres) * abs(yres)) / 1233.48  # 1 AF = 1233.48 m^3
        basinSWE_in = (np.average(value_inches))
        for i in elevationRange:
            # np arrays are mutable, you MUST use copy or eRaster will overwrite the data in grib.elevation_data.
            eRaster = gribObj.elevation_data.copy()
            bottomElevation = i
            topElevation = i + 500
            eRaster[eRaster >= topElevation] = 0
            eRaster[eRaster < bottomElevation] = 0
            eRaster[eRaster > 0] = 1
            #To calculate the average SWE, we have to assume all zero values are outside of the basin, so set all
            #zero values to nan. Note, this would skew areas with zero SWE values in a + direction.
            basinSWE = (value_inches.T * eRaster.T * mask)
            basinSWE[basinSWE == 0] = np.nan
            df.loc[i] = [str(bottomElevation/1000)+'-'+str(topElevation/1000),np.sum(mask * value_AF.T * eRaster.T),np.nanmean(basinSWE)]
            #d[str(bottomElevation/1000)+'-'+str(topElevation/1000)] = [np.sum(mask * value_AF.T * eRaster.T)]
    return df #pd.DataFrame.from_dict(d, orient='index')

def compareDates(mask, gribObj, xres, yres):
    #Numpy arrays are mutable so manipulating gribObj.data will also change grib.data
    grib1 = grib.data.copy()
    grib2 = get_snowdas(grib,inputArgs.date2)
    grib.data = grib1.copy() - grib2.data.copy()
    basinTotal = 0
    if gribObj.units == '[kg/(m^2)]':
        value_mm = grib.data.copy()
        value_m = value_mm * 0.001  # mm to m.
        value_inches = grib.data.copy() * 0.03937  # mm to inches.
        #value_Cubed = totalPrecip_mm * abs(xres) * abs(yres)  # change each grid box to total precip in cubic meters
        value_AF = (value_m * abs(xres) * abs(yres))/ 1233.48  # 1 AF = 1233.48 m^3
        basinSWE_in = (np.average(value_inches))
        basinTotal = [np.sum(mask * value_AF.T)]
    return basinTotal

def elevationRaster():
    # Elevation Source File
    src = gdal.Open(dir_path + '/Shapefiles/elevation.tif', gdalconst.GA_ReadOnly)
    src_proj = src.GetProjection()
    src_geotrans = src.GetGeoTransform()

    # We want a the elevation raster to match the grib raster:
    match_ds = grib.gribAll
    match_proj = match_ds.GetProjection()
    match_geotrans = match_ds.GetGeoTransform()
    width = match_ds.RasterXSize
    height = match_ds.RasterYSize

    # Create a blank destination file that matches our grib raster. Putting in a temp directory:
    dst = gdal.GetDriverByName('GTiff').Create('/vsimem/temp.dat', width, height, 1, gdalconst.GDT_Float32)
    dst.SetGeoTransform(match_geotrans)
    dst.SetProjection(match_proj)

    #Now, insert our elevation source raster into the destination file
    gdal.ReprojectImage(src, dst, src_proj, match_proj, gdalconst.GRA_Bilinear)
    band = dst.GetRasterBand(1)
    elevation_data = BandReadAsArray(band)
    elevation_data[elevation_data < 0] = 0
    grib.elevation_data = elevation_data * 3.28 #convert meters to feet
    del dst  # Flush
    return

def findPointValue(spatialRefProj,xy_source):
    df = pd.read_excel('Snow_Survey_February_2018.xls','Summary',index_col=1)
    #lons, lats = xy_source
    #xx, yy = Proj(spatialRefProj)(lons, lats)
    xx, yy = xy_source

    # Calling a Proj class instance with the arguments lon, lat will convert lon/lat (in degrees) to x/y
    # native map projection coordinates (in meters).
    # If optional keyword 'inverse' is True (default is False), the inverse transformation
    # from x/y to lon/lat is performed.
    model_value = []
    for index, row in df.iterrows():
        lon2x, lat2y = Proj(spatialRefProj)(row['Longitude'],row['Latitude'])
        xIdx = (np.abs(xx - lon2x)).T.argmin()
        yIdx = (np.abs(yy - lat2y)).argmin()
        model_value.append(grib.data.T[xIdx][yIdx])
    df['Model_Value'] = model_value

    # To find the value at a given lat/lon
    # 1) First, take the lat / lon and convert it to X/Y for the given projection.
    #    The reason you do this is because the numpy grid will contain a 2D grid for both the x values and y values
    #    but the x value at point x1,y1 will be same at the point x1,y2 (BUT THE LAT/LON would both be different)
    #    Since you're searching for a unique value (a specific grid box) you want to find the X value you're looking
    #    for and the Y value you're looking for. Once you have your X/Y grid box, you can find the value for that grid box.
    df.to_excel('output.xls', 'Sheet1')
    return df

def makeMap(lons,lats,hr,m,m2,df,deltaDay):
    output_dir = os.path.join(dir_path, 'images', grib.basin, grib.date.strftime("%Y%m%d"))
    imgtype = None
    if imgtype == 'cumulative':
        raster = sum(grib.data[0:grib.hours.index(hr)], axis=0) #cumulative
    else:
        raster = grib.data  # 1 hr forecast (not cumulative)

    if grib.displayunits == 'US' and grib.units == '[kg/(m^2)]':
        raster = raster * 0.03937
        grib.units = 'inches'

    #YOU CAN NOT PUT NAN VAULES IN BEFORE DOING scipy.ndimage.zoom
    raster[raster == 0] = np.nan #this will prevent values of zero from being plotted.

    maxVal = int(np.nanpercentile(raster, 99,  interpolation='linear'))
    minVal = int(np.nanpercentile(raster, 1,  interpolation='linear'))

    im = m.pcolormesh(lons, lats, raster.T, cmap=plt.cm.jet, vmin=minVal, vmax=maxVal) # return 50th percentile, e.g median., latlon=True)
    im2 = m2.pcolormesh(lons, lats, raster.T, cmap=plt.cm.jet, vmin=minVal,vmax=maxVal)  # return 50th percentile, e.g median., latlon=True)
    cb = m.colorbar(mappable=im, location='right', label='SWE (in.)')
    #Show user defined points on map.
    if df != None:
        for index, row in df.iterrows():
            m.plot(row['Longitude'], row['Latitude'], 'ro')
            plt.text(row['Longitude'], row['Latitude'],str(round(row['Model_Value']* 0.03937,1))+' / ' + str(row['SWE']))
            print("Modeled Value: " + str(round(row['Model_Value']* 0.03937,1))+' / Actual Value: ' + str(row['SWE']))


    #plot shapefile
    m.readshapefile(dir_path + '/Shapefiles/' + grib.basin + '/' + grib.basin + '_EPSG4326',
                    grib.basin + '_EPSG4326', linewidth=1)
    m2.readshapefile(dir_path+'/Shapefiles/'+grib.basin+'/'+grib.basin+'_EPSG4326',
                     grib.basin+'_EPSG4326', linewidth=1)
    if grib.basin == 'Hell_Hole':
        m.readshapefile(dir_path + '/Shapefiles/Hell_Hole_SMUD/Hell_Hole_SMUD' + '_EPSG4326',
                    'Hell_Hole_SMUD' + '_EPSG4326', linewidth=1)
        m2.readshapefile(dir_path + '/Shapefiles/Hell_Hole_SMUD/Hell_Hole_SMUD' + '_EPSG4326',
                     'Hell_Hole_SMUD' + '_EPSG4326', linewidth=1)

    # annotate
    m.drawcountries()
    m.drawstates()
    #m.drawrivers()
    m.drawcounties(color='darkred')
    if inputArgs.date2 != None:
        plt.suptitle(grib.basin.replace('_', " ") + ' Difference in SWE between ' + grib.date.strftime("%m/%d/%Y") +
        ' and ' + grib.date2.strftime("%m/%d/%Y") +
                '\n Total Difference in AF (calculated from SWE): ' + str(round(grib.basinTotal[0], 0)) + ' acre feet')
        img = Image.open(output_dir+"/"+grib.date.strftime("%Y%m%d")+"_0_"+grib.basin+'.png')
        w, h = img.size
        draw = ImageDraw.Draw(img)
        font = ImageFont.truetype("micross.ttf", 120) #Avail in C:\\Windows\Fonts
        plus_sign=''
        if grib.basinTotal[0] > 0:
            plus_sign = "+"
        draw.text((1000,h-400),'7 Day Change from ' + grib.date.strftime("%#m/%d") +' to ' +
                  grib.date2.strftime("%#m/%d") + ': ' + plus_sign + str(round(grib.basinTotal[0], 0)) + ' acre feet',(0,0,0), font=font)
        img.save(output_dir+"/"+grib.date.strftime("%Y%m%d")+"_0_"+grib.basin+'.png')
    else:
        plt.suptitle(grib.basin.replace('_'," ") + ' SWE in (inches) for: ' + grib.date.strftime("%m/%d/%Y") +
                  '\n Total AF from SWE: ' + str(round(grib.basinTotal[0],0)) +' acre feet')
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    plt.savefig(output_dir+"/"+grib.date.strftime("%Y%m%d")+"_"+str(-deltaDay)+"_"+grib.basin+'.png',dpi=775)
    print("Saved to " + dir_path+'/images/'+grib.basin+'.png')
    #plt.show()
    plt.close()



def makePlot(elevation_bins,deltaDay):
    output_dir = os.path.join(dir_path,'images',grib.basin,grib.date.strftime("%Y%m%d"))
    sb.set_style('whitegrid')
    chart = sb.barplot(y="TotalAF", x='ElevRange', data=elevation_bins[5:], palette="GnBu_d")
    chart.set_xticklabels(chart.get_xticklabels(), rotation=310)
    chart.set(xlabel="Elevation Range (K ft)", ylabel="Total AF")
    ax2 = chart.twinx()
    ax2.grid(False)
    sb.pointplot(y='AveSWE', x='ElevRange', ax=ax2, data=elevation_bins[5:])
    ax2.set(ylabel='Average SWE (inches)')
    if inputArgs.date2 != None:
        chart.set_title('Change in Total AF by Elevation Range For: ' + grib.basin.replace('_', " ") + ' Basin' +
                        '\n Between ' + grib.date.strftime("%m/%d/%Y") + ' and ' +
                        datetime.datetime.strptime(inputArgs.date2,("%Y%m%d")).strftime("%m/%d/%Y"))
    else:
        chart.set_title('SWE in Total AF by Elevation Range For: ' + grib.basin.replace('_', " ") + ' Basin')
    chart.figure.tight_layout()
    fig = chart.get_figure()
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    fig.savefig(os.path.join(output_dir,grib.date.strftime("%Y%m%d")+"_"+str(-deltaDay)+"_"+grib.basin+'_plot.png'))
    fig.clf() #MUST close figure every time!
    plt.close()
    #chart = sb.factorplot(x='Elevation', y='Total AF', data=elevation_bins,palette="BuPu")
    #chart.show()
    #plt.bar(np.arange(len(elevation_bins)),elevation_bins, color='blue')
    #label = (grib.basinTotal[x] for x in range(0,len(grib.basinTotal)))
    #for i, v in enumerate(elevation_bins):
    #    if v > 0:
    #        plt.text(i, v, str(int(v)))
    #plt.xticks(np.arange(len(elevation_bins)), [str(x)+'-'+str(x+1) + 'K ft' for x in range(0,len(elevation_bins))], rotation=315)
    #plt.title('Total Acre Feet By Elevation')
    #plt.ylabel('Total Acre Feet')
    #plt.xlabel('Elevation (ft)')
    #plt.savefig('BAR_GRAPH.png', dpi=775)
    return

def excel_output(df_elevations):
    efile = os.path.join('G:\Shared Files\Power Marketing\Weather','Daily_Output.xlsx')
    book = load_workbook(efile)
    writer = pd.ExcelWriter(efile, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    dfe = pd.DataFrame([(grib.date.strftime("%m/%d/%Y"),grib.basinTotal[0],df_elevations['AveSWE'].mean())],
                       columns=['Date','TotalAF','AveSWE'])
    dfe.to_excel(writer, sheet_name=grib.basin, header=None, index=None, startrow=writer.sheets[grib.basin].max_row)

    # Write the elevation bands to the elv_bands sheet.
    dfe['Date'].to_excel(writer, sheet_name=grib.basin + '_elv_bands', header=None, index=None,
                                     startrow=writer.sheets[grib.basin + '_elv_bands'].max_row)
    df_bands = df_elevations.set_index(['ElevRange'])
    #The only way to transpose in excel is to drop all columns except the one you're trying to transpose...lame!
    df_bands.drop(['AveSWE'], axis=1, inplace=True)
    df_bands[6:].T.to_excel(writer, sheet_name=grib.basin+'_elv_bands', header=None, index = None,
                                     startrow=writer.sheets[grib.basin+'_elv_bands'].max_row-1, startcol=1)
    writer.save()
    return

def makePNG():
    im = Image.fromarray(((grib.data)/2).astype('uint8'))
    #size = (2181*2), (1205*2)
    #im_resized = im.resize(size, Image.ANTIALIAS)
    im.save('C:/xampp/htdocs/demos/build/static/snowdas.png',"PNG")

if __name__ == "__main__":
    main()